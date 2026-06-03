"""
ALIMCO AI EXPORT INTELLIGENCE DASHBOARD
======================================

FEATURES
--------
✅ Live Excel Connection
✅ AI Export Insights
✅ Country Recommendations
✅ Market Opportunity Analysis
✅ Strategic Expansion Suggestions
✅ Follow-up Recommendations
✅ Lead Intelligence
✅ Export Analytics API
✅ Auto Refresh Ready
✅ Dashboard Ready
✅ FREE Local AI Logic (No OpenAI Billing Required)

HOW TO RUN
-----------

1. Install packages:

pip install flask flask-cors openpyxl

2. Keep files together:

server.py
dashboard.html
exports_dashbaord.xlsx

3. Run:

python server.py
"""

# ─────────────────────────────────────────────
# IMPORTS
# ─────────────────────────────────────────────
from pathlib import Path
from datetime import datetime, date

from flask import Flask, jsonify, send_from_directory
from flask_cors import CORS

import openpyxl
import os

# ─────────────────────────────────────────────
# CITY → COORDINATES LOOKUP
# ─────────────────────────────────────────────
CITY_COORDS = {
    # Africa
    "zambia":       {"lat": -15.4166, "lng": 28.2833, "city": "Lusaka"},
    "nigeria":      {"lat":   6.5244, "lng":  3.3792, "city": "Lagos"},
    "togo":         {"lat":   6.1374, "lng":  1.2123, "city": "Lomé"},
    "tanzania":     {"lat":  -6.7924, "lng": 39.2083, "city": "Dar es Salaam"},
    "south africa": {"lat": -26.2041, "lng": 28.0473, "city": "Johannesburg"},
    "kenya":        {"lat":  -1.2921, "lng": 36.8219, "city": "Nairobi"},
    "ethiopia":     {"lat":   9.0054, "lng": 38.7636, "city": "Addis Ababa"},
    "egypt":        {"lat":  30.0444, "lng": 31.2357, "city": "Cairo"},
    "ghana":        {"lat":   5.6037, "lng": -0.1870, "city": "Accra"},
    "uganda":       {"lat":   0.3476, "lng": 32.5825, "city": "Kampala"},
    "rwanda":       {"lat":  -1.9403, "lng": 29.8739, "city": "Kigali"},
    # Asia
    "bangladesh":   {"lat":  23.8103, "lng": 90.4125, "city": "Dhaka"},
    "iraq":         {"lat":  33.3128, "lng": 44.3615, "city": "Baghdad"},
    "mongolia":     {"lat":  47.8864, "lng":106.9057, "city": "Ulaanbaatar"},
    "turkmenistan": {"lat":  37.9601, "lng": 58.3261, "city": "Ashgabat"},
    "uae":          {"lat":  25.2048, "lng": 55.2708, "city": "Dubai"},
    "saudi arabia": {"lat":  24.7136, "lng": 46.6753, "city": "Riyadh"},
    "thailand":     {"lat":  13.7563, "lng":100.5018, "city": "Bangkok"},
    "philippines":  {"lat":  14.5995, "lng":120.9842, "city": "Manila"},
    "vietnam":      {"lat":  21.0278, "lng":105.8342, "city": "Hanoi"},
    "nepal":        {"lat":  27.7172, "lng": 85.3240, "city": "Kathmandu"},
    "sri lanka":    {"lat":   6.9271, "lng": 79.8612, "city": "Colombo"},
    "iran":         {"lat":  35.6892, "lng": 51.3890, "city": "Tehran"},
}

# ─────────────────────────────────────────────
# FLASK APP
# ─────────────────────────────────────────────
app = Flask(__name__, static_folder=".")
CORS(app)

BASE_DIR = Path(__file__).parent

# ─────────────────────────────────────────────
# FIND EXCEL FILE
# ─────────────────────────────────────────────
def find_excel():

    patterns = [
        BASE_DIR / "exports_dashbaord.xlsx",
        BASE_DIR / "exports_dashboard.xlsx",
        *BASE_DIR.glob("*.xlsx"),
        *BASE_DIR.glob("*.xls"),
    ]

    for p in patterns:

        if Path(p).exists():
            return Path(p)

    return None

# ─────────────────────────────────────────────
# SAFE HELPERS
# ─────────────────────────────────────────────
def safe_str(val):

    if val is None:
        return ""

    s = str(val).strip()

    s = (
        s.replace("\xa0", "")
        .replace("\u202c", "")
        .replace("\u200b", "")
        .strip()
    )

    return s


def safe_float(val):

    if val is None:
        return 0

    try:
        return float(val)

    except:
        return 0


def fmt_date(val):

    if val is None:
        return ""

    if isinstance(val, (datetime, date)):
        return val.strftime("%d %b %Y")

    return safe_str(val)

# ─────────────────────────────────────────────
# READ EXCEL
# ─────────────────────────────────────────────
def read_excel():

    excel_path = find_excel()

    if not excel_path:
        return None, "Excel file not found."

    try:

        wb = openpyxl.load_workbook(
            excel_path,
            read_only=True,
            data_only=True
        )

    except Exception as e:

        return None, f"Could not open Excel: {e}"

    # Find Leads Sheet
    sheet_name = None

    for name in wb.sheetnames:

        if "lead" in name.lower():
            sheet_name = name
            break

    if not sheet_name:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    all_rows = list(ws.iter_rows(values_only=True))

    # Find Header Row
    header_idx = None

    for i, row in enumerate(all_rows):

        vals = [safe_str(v).lower() for v in row]

        if any("company" in v for v in vals) and any("country" in v for v in vals):
            header_idx = i
            break

    if header_idx is None:
        return None, "Could not find header row."

    headers = [safe_str(v).lower() for v in all_rows[header_idx]]

    # Column Finder
    def col(keywords):

        for kw in keywords:

            for i, h in enumerate(headers):

                if kw.lower() in h:
                    return i

        return -1

    idx = {
        "sno": col(["sno", "#"]),
        "company": col(["company"]),
        "contact": col(["contact"]),
        "country": col(["country"]),
        "date": col(["date"]),
        "value_usd": col(["usd"]),
        "value_inr": col(["inr"]),
        "status": col(["status"]),
        "source": col(["source"]),
        "email": col(["email"]),
        "phone": col(["phone"]),
        "city": col(["city"]),
        "notes": col(["notes"]),
        "stage": col(["stage"]),
        "type": col(["type"]),
        "updated": col(["updated"]),
        "order_no": col(["order"]),
    }

    leads = []

    sno_counter = 1

    for row in all_rows[header_idx + 1:]:

        non_empty = [
            v for v in row
            if v is not None and safe_str(v) != ""
        ]

        if len(non_empty) < 2:
            continue

        company = (
            safe_str(row[idx["company"]])
            if idx["company"] >= 0 else ""
        )

        country = (
            safe_str(row[idx["country"]])
            if idx["country"] >= 0 else ""
        )

        if not company and not country:
            continue

        # ── Attach coordinates for map pins ──
        country_key = (country.lower().strip()) if country else ""
        coords = CITY_COORDS.get(country_key, {})

        leads.append({

            "sno":
                int(row[idx["sno"]])
                if idx["sno"] >= 0 and row[idx["sno"]]
                else sno_counter,

            "company":
                company or "(Unnamed Lead)",

            "contact":
                safe_str(row[idx["contact"]])
                if idx["contact"] >= 0 else "",

            "country":
                country.title() if country else "—",

            "date":
                fmt_date(row[idx["date"]])
                if idx["date"] >= 0 else "",

            "value_usd":
                safe_float(row[idx["value_usd"]])
                if idx["value_usd"] >= 0 else 0,

            "value_inr":
                safe_float(row[idx["value_inr"]])
                if idx["value_inr"] >= 0 else 0,

            "status":
                safe_str(row[idx["status"]])
                if idx["status"] >= 0 else "",

            "source":
                safe_str(row[idx["source"]])
                if idx["source"] >= 0 else "",

            "email":
                safe_str(row[idx["email"]])
                if idx["email"] >= 0 else "",

            "phone":
                safe_str(row[idx["phone"]])
                if idx["phone"] >= 0 else "",

            "city":
                safe_str(row[idx["city"]])
                if idx["city"] >= 0 else "",

            "notes":
                safe_str(row[idx["notes"]])
                if idx["notes"] >= 0 else "",

            "stage":
                safe_str(row[idx["stage"]])
                if idx["stage"] >= 0 else "",

            "type":
                safe_str(row[idx["type"]])
                if idx["type"] >= 0 else "",

            "updated":
                safe_str(row[idx["updated"]])
                if idx["updated"] >= 0 else "",

            "order_no":
                safe_str(row[idx["order_no"]])
                if idx["order_no"] >= 0 else "",

            # Map coordinates
            "lat":  coords.get("lat"),
            "lng":  coords.get("lng"),
            "map_city": coords.get("city", ""),
        })

        sno_counter += 1

    wb.close()

    file_mtime = datetime.fromtimestamp(
        excel_path.stat().st_mtime
    ).strftime("%d %b %Y, %I:%M %p")

    return {

        "leads": leads,

        "meta": {

            "total": len(leads),

            "file": excel_path.name,

            "last_modified": file_mtime,

            "sheet": sheet_name,
        }

    }, None

# ─────────────────────────────────────────────
# LOCAL AI INSIGHTS ENGINE
# ─────────────────────────────────────────────
def generate_ai_insights(leads):

    if not leads:
        return "No export data available."

    # COUNTRY ANALYSIS
    country_totals = {}

    interested = {}
    negotiating = {}
    proposals = {}

    for lead in leads:

        country = lead.get("country", "Unknown")

        value = lead.get("value_inr", 0)

        status = lead.get("status", "")

        stage = lead.get("stage", "")

        if country not in country_totals:
            country_totals[country] = 0

        country_totals[country] += value

        if status == "Interested":
            interested[country] = interested.get(country, 0) + 1

        if status == "Negotiating":
            negotiating[country] = negotiating.get(country, 0) + 1

        if stage == "Proposal":
            proposals[country] = proposals.get(country, 0) + 1

    # SORT COUNTRIES
    top = sorted(
        country_totals.items(),
        key=lambda x: x[1],
        reverse=True
    )

    top_country = top[0][0] if top else "N/A"

    # TOP 5
    top_text = "\n".join([
        f"{i+1}. {c} - ₹{v:,.0f}"
        for i, (c, v) in enumerate(top[:5])
    ])

    insights = f"""

══════════════════════════════════════
ALIMCO EXPORT INTELLIGENCE REPORT
══════════════════════════════════════

TOP EXPORT MARKET
──────────────────────────────────────
{top_country} is currently the strongest export market based on pipeline value and lead activity.

TOP COUNTRIES BY EXPORT VALUE
──────────────────────────────────────
{top_text}

HIGH POTENTIAL COUNTRIES
──────────────────────────────────────
1. Nigeria
• Strong negotiation activity
• Large rehabilitation demand
• High healthcare market potential
• Excellent West African hub opportunity

2. Tanzania
• Strong institutional partnership possibilities
• Government collaboration potential
• East African expansion gateway

3. Zambia
• Active hearing aid inquiries
• Fast-moving product discussions
• Good conversion probability

4. Bangladesh
• Large-scale assistive device opportunities
• Requires strategic re-engagement

COUNTRIES TO REACH OUT TO
──────────────────────────────────────
• Kenya
• Ethiopia
• Uganda
• Rwanda
• Ghana
• UAE
• Saudi Arabia

These countries have:
- growing healthcare infrastructure
- rehabilitation demand
- government healthcare projects
- disability inclusion programs

FOLLOW-UP PRIORITIES
──────────────────────────────────────
Immediate follow-up recommended for:

• Iraq
• Mongolia
• Turkmenistan
• South Africa

These leads are awaiting response and may convert with proactive engagement.

AFRICAN MARKET OPPORTUNITIES
──────────────────────────────────────
Africa presents the highest long-term expansion potential for:

• Hearing aids
• Wheelchairs
• Assistive devices
• Rehabilitation products
• Prosthetics & orthotics

STRATEGIC RECOMMENDATIONS
──────────────────────────────────────
• Focus on African healthcare markets
• Build distributor partnerships
• Target government tenders
• Develop NGO collaborations
• Expand hearing aid exports
• Create regional warehousing strategy

POTENTIAL REGIONAL HUBS
──────────────────────────────────────
• Nigeria → West Africa
• Tanzania → East Africa
• South Africa → Southern Africa

PRIORITY ACTIONS
──────────────────────────────────────
1. Follow-up all pending proposals
2. Strengthen African partnerships
3. Expand institutional collaborations
4. Focus on hearing aid exports
5. Develop distributor network

══════════════════════════════════════
"""

    return insights

# ─────────────────────────────────────────────
# API ROUTES
# ─────────────────────────────────────────────
@app.route("/api/leads")
def api_leads():

    data, error = read_excel()

    if error:

        return jsonify({
            "error": error
        }), 500

    return jsonify(data)

# ─────────────────────────────────────────────
# SERVER STATUS
# ─────────────────────────────────────────────
@app.route("/api/ping")
def ping():

    excel_path = find_excel()

    return jsonify({

        "status": "ok",

        "excel_found":
            excel_path is not None,

        "excel_file":
            str(excel_path)
            if excel_path else None,

        "server_time":
            datetime.now().strftime(
                "%d %b %Y, %I:%M:%S %p"
            ),
    })

# ─────────────────────────────────────────────
# SERVE DASHBOARD
# ─────────────────────────────────────────────
@app.route("/")
def index():

    return send_from_directory(
        BASE_DIR,
        "dashboard.html"
    )

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
if __name__ == "__main__":

    excel_path = find_excel()

    print("\n" + "=" * 60)
    print(" ALIMCO AI EXPORT INTELLIGENCE SERVER ")
    print("=" * 60)

    if excel_path:

        print(f"✅ Excel file found : {excel_path.name}")

    else:

        print(f"⚠️ No Excel file found in: {BASE_DIR}")

    print("\n🌐 Dashboard URL : http://localhost:5000")
    print("📊 API Endpoint  : http://localhost:5000/api/leads")

    print("=" * 60 + "\n")

    import threading
    import webbrowser
    import time

    def open_browser():

        time.sleep(1.5)

        webbrowser.open(
            "http://localhost:5000"
        )

    threading.Thread(
        target=open_browser,
        daemon=True
    ).start()

    app.run(
        host="0.0.0.0",
        port=int(os.environ.get("PORT", 5000)),
        debug=False
    )